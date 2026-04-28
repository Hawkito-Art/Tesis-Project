from decimal import Decimal
from io import BytesIO

from django.contrib.auth import get_user_model
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.test import APITestCase

from apps.accounts.models import Role, UserRole
from apps.catalog.models import Entity, Period
from apps.indicators.models import Indicator, IndicatorGroup, IndicatorRecord

User = get_user_model()


class CalculationExportApiTestCase(APITestCase):
    def setUp(self):
        self.url = '/api/exports/xlsx/'

        self.admin_user = User.objects.create_user(
            email='admin-export@test.com',
            password='Passw0rd!123',
            is_staff=True,
        )
        self.analyst_user = User.objects.create_user(
            email='analyst-export@test.com',
            password='Passw0rd!123',
        )
        self.regular_user = User.objects.create_user(
            email='regular-export@test.com',
            password='Passw0rd!123',
        )

        analyst_role, _ = Role.objects.get_or_create(name='analyst', defaults={'description': 'Analyst'})
        UserRole.objects.get_or_create(user=self.analyst_user, role=analyst_role)

        self.entity = Entity.objects.create(code='EXP-001', name='Entidad Export', type='empresa')
        self.other_entity = Entity.objects.create(code='EXP-002', name='Entidad Export 2', type='empresa')
        self.period = Period.objects.create(year=2026, month=4, period_type='mensual')
        self.other_period = Period.objects.create(year=2026, month=5, period_type='mensual')

        self.group = IndicatorGroup.objects.get(group_type='fundamental')
        self.indicator = Indicator.objects.create(
            indicator='EXP_IND_001',
            name='Indicador Export 1',
            unit='MP',
            group=self.group,
        )
        self.other_indicator = Indicator.objects.create(
            indicator='EXP_IND_002',
            name='Indicador Export 2',
            unit='MP',
            group=self.group,
        )

        IndicatorRecord.objects.create(
            entity=self.entity,
            indicator=self.indicator,
            period=self.period,
            variable_name='plan_acumulado',
            value=Decimal('100.5000'),
            source=IndicatorRecord.SOURCE_MANUAL,
        )
        IndicatorRecord.objects.create(
            entity=self.entity,
            indicator=self.indicator,
            period=self.period,
            variable_name='real_acumulado',
            value=Decimal('95.0000'),
            source=IndicatorRecord.SOURCE_IMPORTED,
        )
        IndicatorRecord.objects.create(
            entity=self.other_entity,
            indicator=self.other_indicator,
            period=self.other_period,
            variable_name='plan_acumulado',
            value=Decimal('300.0000'),
            source=IndicatorRecord.SOURCE_CALCULATED,
        )

    def _load_workbook(self, response):
        return load_workbook(filename=BytesIO(response.content))

    def test_export_requires_authentication(self):
        response = self.client.post(self.url, {}, format='json')
        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_export_forbidden_for_user_without_role(self):
        self.client.force_authenticate(user=self.regular_user)
        response = self.client.post(self.url, {}, format='json')
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_export_returns_400_for_invalid_payload(self):
        self.client.force_authenticate(user=self.analyst_user)
        response = self.client.post(self.url, {'entity': 'invalid'}, format='json')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('error', response.data)
        self.assertIn('details', response.data['error'])
        self.assertIn('entity', response.data['error']['details'])

    def test_export_success_returns_xlsx_binary_with_headers(self):
        self.client.force_authenticate(user=self.analyst_user)
        response = self.client.post(self.url, {}, format='json')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('attachment; filename="export_', response['Content-Disposition'])
        self.assertTrue(response.content)

        workbook = self._load_workbook(response)
        worksheet = workbook['indicator_records']
        self.assertEqual(worksheet.max_row, 4)  # header + 3 records

    def test_export_applies_entity_and_period_filters(self):
        self.client.force_authenticate(user=self.admin_user)
        response = self.client.post(
            self.url,
            {'entity': self.entity.id, 'period': self.period.id},
            format='json',
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        workbook = self._load_workbook(response)
        worksheet = workbook['indicator_records']

        rows = list(worksheet.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(rows), 2)
        self.assertTrue(all(row[0] == self.entity.code for row in rows))
        self.assertTrue(all(row[2] == self.period.year for row in rows))
        self.assertTrue(all(row[3] == self.period.month for row in rows))

    def test_export_allows_admin(self):
        self.client.force_authenticate(user=self.admin_user)
        response = self.client.post(self.url, {}, format='json')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
