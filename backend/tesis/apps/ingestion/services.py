from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook
from django.db import transaction
from django.utils import timezone

from apps.indicators.models import Indicator
from apps.indicators.models import IndicatorRecord
from .models import Document, DocumentDetail, ImportJob

CANONICAL_VARIABLE_COLUMN_MAP: dict[int, str] = {
    2: 'plan_anual',
    3: 'ano_anterior',
    4: 'plan_acumulado',
    5: 'real_acumulado',
}

SECTION_ROW_LABELS = {
    'Indicadores Limites',
    'Otros Indicadores',
}


def create_document_and_import_job(*, validated_data: dict[str, Any], uploaded_by) -> tuple[Document, ImportJob]:
    """Crea un documento y su job inicial de importación en estado pendiente."""
    document = Document.objects.create(
        name=validated_data['name'],
        file=validated_data['file'],
        import_type=validated_data['import_type'],
        uploaded_by=uploaded_by,
        status='pendiente',
    )
    import_job = ImportJob.objects.create(
        document=document,
        entity=validated_data.get('entity'),
        period=validated_data.get('period'),
        status='pendiente',
    )
    return document, import_job


def process_import_job_partial(*, import_job: ImportJob) -> ImportJob:
    """Procesa importación parcial: guarda detalles por fila y actualiza métricas."""
    import_job.status = 'en_progreso'
    import_job.started_at = timezone.now()
    import_job.save(update_fields=['status', 'started_at'])

    document = import_job.document
    document.status = 'procesado'
    document.save(update_fields=['status'])

    parsed_rows = parse_document_indicator_rows(document=import_job.document)

    details_to_create: list[DocumentDetail] = []
    error_messages: list[str] = []
    error_rows = 0

    for row in parsed_rows:
        if not row['is_valid']:
            error_rows += 1
            error_messages.append(f"fila {row['row_number']}: {row['error_message']}")

        details_to_create.append(
            DocumentDetail(
                import_job=import_job,
                row_number=row['row_number'],
                raw_data={
                    'indicator_name': row['indicator_name'],
                    'raw_values': row['raw_values'],
                    'parsed_values': {
                        key: (str(value) if value is not None else None)
                        for key, value in row['parsed_values'].items()
                    },
                },
                is_valid=row['is_valid'],
                error_message=row['error_message'],
            )
        )

    with transaction.atomic():
        DocumentDetail.objects.filter(import_job=import_job).delete()
        if details_to_create:
            DocumentDetail.objects.bulk_create(details_to_create)

        import_job.total_rows = len(parsed_rows)
        import_job.processed_rows = len(parsed_rows)
        import_job.error_rows = error_rows
        import_job.error_log = '\n'.join(error_messages)
        import_job.status = 'completado'
        import_job.finished_at = timezone.now()
        import_job.save(
            update_fields=[
                'total_rows',
                'processed_rows',
                'error_rows',
                'error_log',
                'status',
                'finished_at',
            ]
        )

    return import_job


def upsert_indicator_records_from_import_job(
    *,
    import_job: ImportJob,
    entity,
    period,
) -> dict[str, int]:
    """Persiste variables base válidas como upsert en IndicatorRecord."""
    details = DocumentDetail.objects.filter(import_job=import_job, is_valid=True).order_by('row_number')
    created_count = 0
    updated_count = 0

    for detail in details:
        indicator_name = detail.raw_data.get('indicator_name')
        indicator = Indicator.objects.filter(name=indicator_name, is_active=True).first()
        if indicator is None:
            continue

        parsed_values = detail.raw_data.get('parsed_values', {})
        for variable_name in ('plan_anual', 'ano_anterior', 'plan_acumulado', 'real_acumulado'):
            raw_value = parsed_values.get(variable_name)
            value = _to_decimal_or_none(raw_value)

            record, created = IndicatorRecord.objects.update_or_create(
                entity=entity,
                indicator=indicator,
                period=period,
                variable_name=variable_name,
                defaults={
                    'value': value,
                    'source': IndicatorRecord.SOURCE_IMPORTED,
                    'import_job': import_job,
                    'calculation': None,
                },
            )
            if created:
                created_count += 1
            else:
                updated_count += 1

    return {
        'created': created_count,
        'updated': updated_count,
        'total': created_count + updated_count,
    }


def parse_document_indicator_rows(*, document: Document) -> list[dict[str, Any]]:
    """Parsea workbook .xlsx y devuelve filas candidatas para staging de indicadores."""
    workbook = load_workbook(document.file.path, data_only=True)
    worksheet = workbook.active
    header_row = _find_header_row(worksheet)
    indicators_by_name = {
        indicator.name: indicator
        for indicator in Indicator.objects.filter(is_active=True)
    }

    parsed_rows: list[dict[str, Any]] = []
    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        indicator_name = _to_clean_str(worksheet.cell(row=row_idx, column=1).value)
        if _should_skip_row(indicator_name):
            continue

        raw_values: dict[str, Any] = {}
        parsed_values: dict[str, Decimal | None] = {}
        for column_index, variable_name in CANONICAL_VARIABLE_COLUMN_MAP.items():
            raw_value = worksheet.cell(row=row_idx, column=column_index + 1).value
            raw_values[variable_name] = raw_value
            parsed_values[variable_name] = _to_decimal_or_none(raw_value)

        indicator = indicators_by_name.get(indicator_name)
        if indicator is None:
            parsed_rows.append(
                {
                    'row_number': row_idx,
                    'indicator_name': indicator_name,
                    'indicator_id': None,
                    'is_valid': False,
                    'error_message': f'Indicador no encontrado por nombre exacto: {indicator_name}',
                    'raw_values': raw_values,
                    'parsed_values': parsed_values,
                }
            )
            continue

        parsed_rows.append(
            {
                'row_number': row_idx,
                'indicator_name': indicator_name,
                'indicator_id': indicator.id,
                'is_valid': True,
                'error_message': '',
                'raw_values': raw_values,
                'parsed_values': parsed_values,
            }
        )

    return parsed_rows


def _find_header_row(worksheet) -> int:
    for row_idx in range(1, worksheet.max_row + 1):
        first_cell = _to_clean_str(worksheet.cell(row=row_idx, column=1).value)
        if first_cell == 'INDICADORES':
            return row_idx
    raise ValueError('No se encontro la fila de encabezado INDICADORES en el documento.')


def _should_skip_row(indicator_name: str) -> bool:
    if not indicator_name:
        return True
    if indicator_name in SECTION_ROW_LABELS:
        return True
    if indicator_name == 'INDICADORES':
        return True
    if indicator_name.startswith('Aprobado'):
        return True
    return False


def _to_clean_str(value: Any) -> str:
    if value is None:
        return ''
    return str(value).strip()


def _to_decimal_or_none(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))

    normalized = str(value).strip()
    if not normalized:
        return None
    normalized = normalized.replace(',', '')
    try:
        return Decimal(normalized)
    except InvalidOperation:
        return None
