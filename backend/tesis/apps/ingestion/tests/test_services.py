import shutil
import tempfile

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from openpyxl import Workbook

from apps.indicators.models import Indicator, IndicatorGroup
from apps.ingestion.models import Document
from apps.ingestion.services import parse_document_indicator_rows


class IngestionParserServiceTestCase(TestCase):
    def setUp(self):
        self.temp_media_root = tempfile.mkdtemp(prefix='tesis-test-media-')
        self.group = IndicatorGroup.objects.create(name='Fundamental', group_type='fundamental')
        self.ventas = Indicator.objects.create(
            indicator='VENTAS_TOT',
            name='Ventas Totales',
            unit='MP',
            group=self.group,
        )

    def tearDown(self):
        shutil.rmtree(self.temp_media_root, ignore_errors=True)

    @override_settings(MEDIA_ROOT='/tmp')
    def test_override_settings_guard(self):
        # Guard to keep Django from caching override decorators across tests.
        self.assertTrue(True)

    def _create_document_from_workbook(self, workbook: Workbook) -> Document:
        with override_settings(MEDIA_ROOT=self.temp_media_root):
            from io import BytesIO

            buffer = BytesIO()
            workbook.save(buffer)
            buffer.seek(0)
            uploaded = SimpleUploadedFile(
                'indicadores.xlsx',
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            return Document.objects.create(name='Carga test', file=uploaded)

    def test_parse_document_maps_canonical_columns_and_exact_indicator(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Entidad 1'
        sheet.append(['Titulo'])
        sheet.append([
            'INDICADORES',
            'U:M',
            'AÑO',
            'Año Anter',
            'PLAN',
            'REAL',
            'R/P',
            'R/AA',
            'Estimado mes',
            'Estimado cierre',
        ])
        sheet.append(['Ventas Totales', 'MP', 145881.2, 100209.3, 145881.2, 121785.0, 83.5, 102.3, 10764.9, 145881.2])

        document = self._create_document_from_workbook(workbook)
        with override_settings(MEDIA_ROOT=self.temp_media_root):
            parsed_rows = parse_document_indicator_rows(document=document)

        self.assertEqual(len(parsed_rows), 1)
        row = parsed_rows[0]
        self.assertTrue(row['is_valid'])
        self.assertEqual(row['indicator_id'], self.ventas.id)
        self.assertEqual(row['indicator_name'], 'Ventas Totales')
        self.assertIn('plan_anual', row['parsed_values'])
        self.assertIn('ano_anterior', row['parsed_values'])
        self.assertIn('plan_acumulado', row['parsed_values'])
        self.assertIn('real_acumulado', row['parsed_values'])

    def test_parse_document_marks_unknown_indicator_as_invalid(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['INDICADORES', 'U:M', 'AÑO', 'Año Anter', 'PLAN', 'REAL', 'R/P', 'R/AA', 'Estimado mes', 'Estimado cierre'])
        sheet.append(['Indicador Desconocido', 'MP', 100, 90, 100, 80, 80, 88, 88, 100])

        document = self._create_document_from_workbook(workbook)
        with override_settings(MEDIA_ROOT=self.temp_media_root):
            parsed_rows = parse_document_indicator_rows(document=document)

        self.assertEqual(len(parsed_rows), 1)
        row = parsed_rows[0]
        self.assertFalse(row['is_valid'])
        self.assertIsNone(row['indicator_id'])
        self.assertIn('Indicador no encontrado por nombre exacto', row['error_message'])

    def test_parse_document_skips_section_rows(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['INDICADORES', 'U:M', 'AÑO', 'Año Anter', 'PLAN', 'REAL', 'R/P', 'R/AA', 'Estimado mes', 'Estimado cierre'])
        sheet.append(['Indicadores Limites', '', '', '', '', '', '', '', '', ''])
        sheet.append(['Otros Indicadores', '', '', '', '', '', '', '', '', ''])
        sheet.append(['Ventas Totales', 'MP', 145881.2, 100209.3, 145881.2, 121785.0, 83.5, 102.3, 10764.9, 145881.2])

        document = self._create_document_from_workbook(workbook)
        with override_settings(MEDIA_ROOT=self.temp_media_root):
            parsed_rows = parse_document_indicator_rows(document=document)

        self.assertEqual(len(parsed_rows), 1)
        self.assertEqual(parsed_rows[0]['indicator_name'], 'Ventas Totales')

    def test_parse_document_normalizes_thousand_separator_values(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['INDICADORES', 'U:M', 'AÑO', 'Año Anter', 'PLAN', 'REAL', 'R/P', 'R/AA', 'Estimado mes', 'Estimado cierre'])
        sheet.append(['Ventas Totales', 'MP', '145,881.2', '100,209.3', '145,881.2', '121,785.0', 83.5, 102.3, 10764.9, 145881.2])

        document = self._create_document_from_workbook(workbook)
        with override_settings(MEDIA_ROOT=self.temp_media_root):
            parsed_rows = parse_document_indicator_rows(document=document)

        self.assertEqual(len(parsed_rows), 1)
        parsed = parsed_rows[0]['parsed_values']
        self.assertEqual(parsed['plan_anual'], self._d('145881.2'))
        self.assertEqual(parsed['ano_anterior'], self._d('100209.3'))
        self.assertEqual(parsed['real_acumulado'], self._d('121785.0'))

    def test_parse_document_is_case_sensitive_for_exact_name_mapping(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['INDICADORES', 'U:M', 'AÑO', 'Año Anter', 'PLAN', 'REAL', 'R/P', 'R/AA', 'Estimado mes', 'Estimado cierre'])
        sheet.append(['ventas totales', 'MP', 145881.2, 100209.3, 145881.2, 121785.0, 83.5, 102.3, 10764.9, 145881.2])

        document = self._create_document_from_workbook(workbook)
        with override_settings(MEDIA_ROOT=self.temp_media_root):
            parsed_rows = parse_document_indicator_rows(document=document)

        self.assertEqual(len(parsed_rows), 1)
        self.assertFalse(parsed_rows[0]['is_valid'])
        self.assertIn('Indicador no encontrado por nombre exacto', parsed_rows[0]['error_message'])

    @staticmethod
    def _d(value: str):
        from decimal import Decimal

        return Decimal(value)
